import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = [
    ("nsg-d-weu-idp-001", "Network Security Group", "nsg-idpquote-d-weu-001"),
    ("rsv-d-weu-idp-001", "Recovery Services Vault", "rsv-idpquote-d-weu-001"),
    ("rt-d-weu-idp-afw", "Route Table", "rt-idpquote-d-weu-001"),
    ("vnet-d-weu-idp-001", "Virtual Network", "vnet-idpquote-d-weu-001"),
    ("stdeweuidpquote", "Storage Account", "stdeweuidpquote001"),
    ("Failure Anomalies - idp-quote-dev-ai", "Smart detector alert rule", "alert-d-global-idpquote-001"),
    ("idp-quote-dev-ai", "Application Insights", "appi-d-weu-idpquote-001"),
    ("idp-quote-dev-app", "App Service", "app-d-weu-idpquote-001"),
    ("idp-quote-dev-law", "Log Analytics Workspace", "law-d-weu-idpquote-001"),
    ("idp-quote-dev-openai", "Azure OpenAI", "openai-d-weu-idpquote-001"),
    ("idp-quote-dev-pgsql", "Azure Database for PostgreSQL", "psql-d-weu-idpquote-001"),
    ("idp-quote-dev-plan", "App Service Plan", "asp-d-weu-idpquote-001"),
    ("idpquotedkv", "Key Vault", "kv-d-weu-idpquote-001"),
    ("idpquotedvesa", "Storage Account", "stidpquotedweu001"),
    ("ai-search-dev-weu-idpquote", "Search Service", "srch-d-weu-idpquote-001"),
    ("app-eriks-trackline-dev", "App Service", "app-d-weu-idpquote-001"),
    ("asp-dev-weu-idpquote", "App Service Plan", "asp-d-weu-idpquote-001"),
    ("kv-dev-weu-idpquote", "Key Vault", "kv-d-weu-idpquote-001"),
    ("postgresdb-idpquote-trackline-002", "PostgreSQL Database", "psql-d-weu-idpquote-002"),
    ("rg-d-weu-idpapp-001-vnet", "Virtual Network", "vnet-d-weu-idpquote-001"),
    ("stdeweuidpquote", "Storage Account", "stdweuidpquote001"),
    ("Failure Anomalies - idp-quote-test-ai", "Smart detector alert rule", "alert-t-global-idpquote-001"),
    ("idp-quote-test-ai", "Application Insights", "appi-t-weu-idpquote-001"),
    ("idp-quote-test-app", "App Service", "app-t-weu-idpquote-001"),
    ("idp-quote-test-law", "Log Analytics Workspace", "law-t-weu-idpquote-001"),
    ("idp-quote-test-openai", "Azure OpenAI", "openai-t-weu-idpquote-001"),
    ("idp-quote-test-pgsql", "Azure Database for PostgreSQL", "psql-t-weu-idpquote-001"),
    ("idp-quote-test-plan", "App Service Plan", "asp-t-weu-idpquote-001"),
    ("idpquotetestkv", "Key Vault", "kv-t-weu-idpquote-001"),
    ("idpquotetestsa", "Storage Account", "stidpquotetestweu001"),
    ("foundry-prd-idpquote", "Foundry", "foundry-p-weu-idpquote-001"),
    ("proj-foundry-prd-idpquote", "Foundry Project", "foundryproj-p-weu-idpquote-001"),
    ("postgresdb-idpquote-trackline-001.private.postgres.database.azure.com", "Private DNS Zone", "pdns-p-weu-idpquote-001"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resource Renaming"

header_fill = PatternFill("solid", fgColor="0078D4")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill("solid", fgColor="F0F7FF")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Current Name", "Resource Type", "Suggested"]
widths = [70, 35, 35]

for col, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
    ws.column_dimensions[c.column_letter].width = w

ws.row_dimensions[1].height = 20

for i, (current, rtype, suggested) in enumerate(rows, 2):
    fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for col, val in enumerate([current, rtype, suggested], 1):
        c = ws.cell(row=i, column=col, value=val)
        c.fill = fill
        c.border = border
        c.alignment = Alignment(vertical="center")

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:C34"

wb.save("Azure_Resource_Renaming.xlsx")
print("Done - Azure_Resource_Renaming.xlsx created")